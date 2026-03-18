from __future__ import annotations

import argparse
import json
import math
import sys
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as error:  # pragma: no cover - import guard for runtime
    raise SystemExit(
        "Missing dependency: openpyxl. Install it in the active Python environment first."
    ) from error


UTC = timezone.utc
LOCAL_TZ = timezone(timedelta(hours=8))
DEFAULT_SOURCE_FILES = [
    Path(r"d:\OneDrive\文档\2024data.xlsx"),
    Path(r"d:\OneDrive\文档\20251data.xlsx"),
    Path(r"d:\OneDrive\文档\20252data.xlsx"),
    Path(r"d:\OneDrive\文档\2026data.xlsx"),
]
DEFAULT_OUTPUT_DIR = Path(r"f:\code\Order\out\imports")
OUTPUT_JSON_NAME = "order-external-import-all.json"
OUTPUT_SUMMARY_NAME = "order-external-import-all-summary.json"
EXPECTED_TOTAL_RECORDS = 9688
START_TIME_TOLERANCE_MS = 2000
MIN_DURATION_MS = 1
REQUIRED_HEADERS = {
    "用时",
    "当前时间",
    "上个项目",
    "shijianchuocha",
    "time",
    "项目",
}
OPTIONAL_START_HEADER = "记录开始时间"


@dataclass
class WorkbookRow:
    source_workbook: str
    source_sheet: str
    source_row: int
    source_current_text: str
    source_previous_project: str
    source_next_project: str
    source_duration_ms: int
    source_anchor_end_ms: int
    source_explicit_start_ms: int | None
    project_name: str
    start_ms: int
    end_ms: int

    def to_external_record(self) -> dict[str, Any]:
        return {
            "projectName": self.project_name,
            "startTime": format_utc_ms(self.start_ms),
            "endTime": format_utc_ms(self.end_ms),
            "durationMs": self.end_ms - self.start_ms,
            "sourceWorkbook": self.source_workbook,
            "sourceSheet": self.source_sheet,
            "sourceRow": self.source_row,
            "sourceCurrentText": self.source_current_text,
            "sourcePreviousProject": self.source_previous_project,
            "sourceNextProject": self.source_next_project,
            "sourceDurationMs": self.source_duration_ms,
        }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a root-array external import JSON from exported Excel time-chain files."
    )
    parser.add_argument(
        "--source",
        dest="sources",
        action="append",
        default=[],
        help="Excel source path. Repeat to override defaults.",
    )
    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help=f"Output directory. Default: {DEFAULT_OUTPUT_DIR}",
    )
    return parser.parse_args()


def ensure_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_int_ms(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if not math.isfinite(value):
            return None
        return int(round(float(value)))
    text = ensure_text(value)
    if not text:
        return None
    try:
        parsed = float(text)
    except ValueError:
        return None
    if not math.isfinite(parsed):
        return None
    return int(round(parsed))


def parse_iso_utc_ms(value: Any) -> int:
    text = ensure_text(value)
    if not text:
        raise ValueError("Missing ISO time value")
    normalized = text.replace("Z", "+00:00")
    parsed = datetime.fromisoformat(normalized)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=UTC)
    else:
        parsed = parsed.astimezone(UTC)
    return int(round(parsed.timestamp() * 1000))


def format_utc_ms(value: int) -> str:
    return (
        datetime.fromtimestamp(value / 1000, tz=UTC)
        .isoformat(timespec="milliseconds")
        .replace("+00:00", "Z")
    )


def format_local_ms(value: int) -> str:
    return datetime.fromtimestamp(value / 1000, tz=UTC).astimezone(LOCAL_TZ).isoformat(
        timespec="milliseconds"
    )


def read_header_map(sheet: Any) -> dict[str, int]:
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if first_row is None:
        raise ValueError("Worksheet is empty")
    headers = {ensure_text(cell): index for index, cell in enumerate(first_row)}
    missing = sorted(REQUIRED_HEADERS - set(headers))
    if missing:
        raise ValueError(f"Missing required headers: {', '.join(missing)}")
    return headers


def get_row_value(row: tuple[Any, ...], header_map: dict[str, int], header_name: str) -> Any:
    index = header_map.get(header_name)
    if index is None or index >= len(row):
        return None
    return row[index]


def load_workbook_rows(path: Path) -> tuple[list[WorkbookRow], dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if not workbook.sheetnames:
        raise ValueError(f"No worksheets found in {path}")
    sheet = workbook[workbook.sheetnames[0]]
    header_map = read_header_map(sheet)
    rows: list[WorkbookRow] = []
    chain_breaks: list[dict[str, Any]] = []
    previous_start_ms: int | None = None
    previous_next_project: str | None = None
    explicit_start_diffs: list[int] = []
    zero_duration_adjusted_count = 0
    correction_row_count = 0

    for source_row_index, raw_row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        source_previous_project = ensure_text(get_row_value(raw_row, header_map, "上个项目"))
        source_next_project = ensure_text(get_row_value(raw_row, header_map, "项目"))
        project_name = source_previous_project
        if not project_name:
            raise ValueError(f"{path.name} row {source_row_index}: project name is empty")

        source_current_text = ensure_text(get_row_value(raw_row, header_map, "当前时间"))
        source_duration_ms = parse_int_ms(get_row_value(raw_row, header_map, "shijianchuocha"))
        if source_duration_ms is None or source_duration_ms < 0:
            raise ValueError(
                f"{path.name} row {source_row_index}: invalid shijianchuocha value"
            )
        effective_duration_ms = source_duration_ms
        if effective_duration_ms == 0:
            effective_duration_ms = MIN_DURATION_MS
            zero_duration_adjusted_count += 1
        if source_current_text == "修正记录":
            correction_row_count += 1

        source_anchor_end_ms = parse_iso_utc_ms(get_row_value(raw_row, header_map, "time"))
        explicit_start_ms = parse_int_ms(
            get_row_value(raw_row, header_map, OPTIONAL_START_HEADER)
        )

        end_ms = source_anchor_end_ms if previous_start_ms is None else previous_start_ms
        start_ms = (
            explicit_start_ms
            if explicit_start_ms is not None
            else end_ms - effective_duration_ms
        )
        if start_ms >= end_ms:
            raise ValueError(
                f"{path.name} row {source_row_index}: startTime must be earlier than endTime"
            )

        row = WorkbookRow(
            source_workbook=path.name,
            source_sheet=sheet.title,
            source_row=source_row_index,
            source_current_text=source_current_text,
            source_previous_project=source_previous_project,
            source_next_project=source_next_project,
            source_duration_ms=source_duration_ms,
            source_anchor_end_ms=source_anchor_end_ms,
            source_explicit_start_ms=explicit_start_ms,
            project_name=project_name,
            start_ms=start_ms,
            end_ms=end_ms,
        )
        rows.append(row)

        if explicit_start_ms is not None:
            explicit_start_diffs.append(abs(start_ms - explicit_start_ms))

        if previous_next_project is not None and previous_next_project != project_name:
            chain_breaks.append(
                {
                    "previousRow": source_row_index - 1,
                    "expectedCurrentProject": previous_next_project,
                    "actualCurrentProject": project_name,
                }
            )

        previous_start_ms = start_ms
        previous_next_project = source_next_project

    if not rows:
        raise ValueError(f"{path.name}: no data rows found")

    rows_in_chronological_order = sorted(rows, key=lambda item: (item.start_ms, item.end_ms, item.source_row))
    project_names = sorted({row.project_name for row in rows})
    source_duration_mismatch_count = sum(
        1
        for row in rows
        if abs((row.end_ms - row.start_ms) - row.source_duration_ms) > 0
    )
    per_workbook_summary = {
        "sourceWorkbook": path.name,
        "sourceSheet": sheet.title,
        "recordCount": len(rows),
        "projectCount": len(project_names),
        "projectNamesSample": project_names[:20],
        "earliestStartTimeUtc": format_utc_ms(rows_in_chronological_order[0].start_ms),
        "latestEndTimeUtc": format_utc_ms(rows_in_chronological_order[-1].end_ms),
        "earliestStartTimeLocal": format_local_ms(rows_in_chronological_order[0].start_ms),
        "latestEndTimeLocal": format_local_ms(rows_in_chronological_order[-1].end_ms),
        "topRowAnchorUtc": format_utc_ms(rows[0].source_anchor_end_ms),
        "mappingSemantics": {
            "recordProjectField": "上个项目",
            "nextProjectField": "项目",
            "durationField": "shijianchuocha",
            "rawTimeField": "time",
            "correctionRowFlag": "当前时间 == 修正记录",
        },
        "explicitStartCount": sum(
            1 for row in rows if row.source_explicit_start_ms is not None
        ),
        "explicitStartDiffMs": build_numeric_stats(explicit_start_diffs),
        "zeroDurationAdjustedCount": zero_duration_adjusted_count,
        "correctionRowCount": correction_row_count,
        "sourceDurationMismatchCount": source_duration_mismatch_count,
        "chainMonotonic": validate_monotonic_chain(rows),
        "projectTransitionChainBreakCount": len(chain_breaks),
        "projectTransitionChainBreakSamples": chain_breaks[:10],
    }
    return rows_in_chronological_order, per_workbook_summary


def build_numeric_stats(values: list[int]) -> dict[str, Any]:
    if not values:
        return {
            "count": 0,
            "max": None,
            "median": None,
            "p95": None,
            "withinToleranceCount": 0,
            "toleranceMs": START_TIME_TOLERANCE_MS,
        }
    ordered = sorted(values)
    p95_index = min(len(ordered) - 1, max(0, math.ceil(len(ordered) * 0.95) - 1))
    return {
        "count": len(ordered),
        "max": ordered[-1],
        "median": ordered[len(ordered) // 2],
        "p95": ordered[p95_index],
        "withinToleranceCount": sum(1 for value in ordered if value <= START_TIME_TOLERANCE_MS),
        "toleranceMs": START_TIME_TOLERANCE_MS,
    }


def validate_monotonic_chain(rows_in_source_order_desc: list[WorkbookRow]) -> bool:
    if not rows_in_source_order_desc:
        return True
    for index, row in enumerate(rows_in_source_order_desc):
        if row.start_ms >= row.end_ms:
            return False
        if index + 1 >= len(rows_in_source_order_desc):
            continue
        next_row = rows_in_source_order_desc[index + 1]
        if row.start_ms != next_row.end_ms:
            return False
    return True


def build_boundary_offsets(workbook_summaries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ordered_summaries = sorted(
        workbook_summaries,
        key=lambda item: parse_iso_utc_ms(item["earliestStartTimeUtc"]),
    )
    boundaries: list[dict[str, Any]] = []
    for index in range(len(ordered_summaries) - 1):
        older = ordered_summaries[index]
        newer = ordered_summaries[index + 1]
        older_latest_end_ms = parse_iso_utc_ms(older["latestEndTimeUtc"])
        newer_earliest_start_ms = parse_iso_utc_ms(newer["earliestStartTimeUtc"])
        offset_ms = newer_earliest_start_ms - older_latest_end_ms
        boundaries.append(
            {
                "olderWorkbook": older["sourceWorkbook"],
                "newerWorkbook": newer["sourceWorkbook"],
                "olderLatestEndUtc": older["latestEndTimeUtc"],
                "newerEarliestStartUtc": newer["earliestStartTimeUtc"],
                "offsetMs": offset_ms,
                "withinAssumedBoundaryTolerance": abs(offset_ms) <= 250,
            }
        )
    return boundaries


def build_validation_summary(
    records: list[WorkbookRow],
    workbook_summaries: list[dict[str, Any]],
) -> dict[str, Any]:
    total_count_ok = len(records) == EXPECTED_TOTAL_RECORDS
    non_empty_project_names = all(row.project_name for row in records)
    valid_time_ranges = all(row.start_ms < row.end_ms for row in records)
    ascending_order = all(
        records[index].start_ms <= records[index + 1].start_ms
        for index in range(len(records) - 1)
    )
    start_accuracy = next(
        (
            summary["explicitStartDiffMs"]
            for summary in workbook_summaries
            if summary["sourceWorkbook"].lower() == "20252data.xlsx"
        ),
        build_numeric_stats([]),
    )

    return {
        "expectedTotalRecords": EXPECTED_TOTAL_RECORDS,
        "actualTotalRecords": len(records),
        "recordCountMatchesExpectation": total_count_ok,
        "allProjectNamesNonEmpty": non_empty_project_names,
        "allTimeRangesValid": valid_time_ranges,
        "globalStartTimeAscending": ascending_order,
        "explicitStartValidation20252": {
            **start_accuracy,
            "allWithinTolerance": start_accuracy["count"] == start_accuracy["withinToleranceCount"],
        },
        "allWorkbookChainsMonotonic": all(
            summary["chainMonotonic"] for summary in workbook_summaries
        ),
    }


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    args = parse_args()
    source_paths = [Path(value) for value in args.sources] if args.sources else DEFAULT_SOURCE_FILES
    output_dir = Path(args.output_dir)
    all_records: list[WorkbookRow] = []
    workbook_summaries: list[dict[str, Any]] = []

    for source_path in source_paths:
        rows, workbook_summary = load_workbook_rows(source_path)
        all_records.extend(rows)
        workbook_summaries.append(workbook_summary)

    all_records.sort(key=lambda item: (item.start_ms, item.end_ms, item.project_name, item.source_workbook, item.source_row))
    workbook_summaries.sort(key=lambda item: parse_iso_utc_ms(item["earliestStartTimeUtc"]))
    boundary_offsets = build_boundary_offsets(workbook_summaries)
    validation = build_validation_summary(all_records, workbook_summaries)

    if not validation["recordCountMatchesExpectation"]:
        raise SystemExit(
            f"Expected {EXPECTED_TOTAL_RECORDS} records but built {len(all_records)} records."
        )
    if not validation["allProjectNamesNonEmpty"] or not validation["allTimeRangesValid"]:
        raise SystemExit("Validation failed: empty project names or invalid time ranges detected.")
    if not validation["explicitStartValidation20252"]["allWithinTolerance"]:
        raise SystemExit("Validation failed: 20252 explicit start times exceed tolerance.")

    records_payload = [row.to_external_record() for row in all_records]
    summary_payload = {
        "outputJson": str((output_dir / OUTPUT_JSON_NAME).resolve()),
        "outputSummary": str((output_dir / OUTPUT_SUMMARY_NAME).resolve()),
        "generatedAtUtc": datetime.now(tz=UTC).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "sourceFiles": [str(path) for path in source_paths],
        "recordCount": len(all_records),
        "timeRangeUtc": {
            "earliestStart": format_utc_ms(all_records[0].start_ms),
            "latestEnd": format_utc_ms(all_records[-1].end_ms),
        },
        "timeRangeLocal": {
            "earliestStart": format_local_ms(all_records[0].start_ms),
            "latestEnd": format_local_ms(all_records[-1].end_ms),
        },
        "projectCount": len({row.project_name for row in all_records}),
        "workbooks": workbook_summaries,
        "boundaryOffsets": boundary_offsets,
        "validation": validation,
    }

    write_json(output_dir / OUTPUT_JSON_NAME, records_payload)
    write_json(output_dir / OUTPUT_SUMMARY_NAME, summary_payload)

    print(f"Wrote {len(records_payload)} records to {output_dir / OUTPUT_JSON_NAME}")
    print(f"Wrote summary to {output_dir / OUTPUT_SUMMARY_NAME}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
